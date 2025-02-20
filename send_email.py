import smtplib
import openpyxl  # Use openpyxl instead of xlrd
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# Email Configuration
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
EMAIL_ADDRESS = "ambivertamit@gmail.com"  # Replace with your email
EMAIL_PASSWORD = "***"  # Use an App Password

# File Paths
EXCEL_FILE = "emails.xlsx"  # Change if needed
PDF_ATTACHMENT = "amit_resume.pdf"  # Change if needed

def read_emails_from_excel():
    """Reads email addresses from an XLSX file using openpyxl."""
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active  # Get the first sheet
        emails = [sheet.cell(row=row, column=1).value for row in range(2, sheet.max_row + 1) if sheet.cell(row=row, column=1).value]
        return emails
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return []

def send_email(recipient_email):
    """Sends an email with an attachment to a single recipient."""
    msg = MIMEMultipart()
    msg["From"] = EMAIL_ADDRESS
    msg["To"] = recipient_email
    msg["Subject"] = "Application for SDE-I Position"

    email_body = """\
Dear Hiring Manager,
I hope you are doing well.

I am excited to apply for the Full-Stack Developer role at your organization and have attached my resume for your review. With 2 years of experience in a product-based startup, I have honed my skills in full-stack development, making me a strong candidate for this position.

Here are my details for your reference:

Total Experience: 2 years (Product-based startup)
Notice Period: Immediate joiner
Current Location: Pune (Open to relocation)
I would appreciate the opportunity to discuss how my skills and experience align with your team’s needs. I am looking forward to your response.

Best regards,
Amit K.
📞 8094071021
📧 ambivertamit@gmail.com
    
    """
    msg.attach(MIMEText(email_body, "plain"))

    # Attach PDF
    try:
        with open(PDF_ATTACHMENT, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename={PDF_ATTACHMENT}")
            msg.attach(part)
    except Exception as e:
        print(f"Failed to attach PDF for {recipient_email}: {e}")
        return

    # Send Email
    try:
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        server.sendmail(EMAIL_ADDRESS, recipient_email, msg.as_string())
        server.quit()
        print(f"Email sent to: {recipient_email}")
    except Exception as e:
        print(f"Failed to send email to {recipient_email}: {e}")

def send_bulk_emails():
    """Reads emails from Excel and sends emails to all recipients."""
    emails = read_emails_from_excel()
    for email in emails:
        send_email(email)

# Run only if executed directly
if __name__ == "__main__":
    send_bulk_emails()
